from pathlib import Path
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog
from database import (
    get_all_enabled_car_nos,
    get_destination_whitelist,
    set_destination_whitelist,
    is_car_allowed_for_destination,
    complete_vehicle_out,
    get_trip_records_for_export,
)

import openpyxl
import flet as ft
import threading
import time
import asyncio
from database import (
    init_db,
    seed_sample_data,
    get_vehicles,
    get_destinations,
    get_loading_levels,
    add_vehicle,
    update_vehicle,
    delete_vehicle,
    add_destination,
    update_destination,
    delete_destination,
    add_loading_level,
    update_loading_level,
    delete_loading_level,
    update_vehicle_status,
    update_vehicle_out,
    update_vehicle_loading,
    update_vehicle_repair,

    seed_default_admin,
    authenticate_user,
    get_users,
    add_user,
    update_user,
    delete_user,
    get_destination_minutes,
)


STATUS_LIST = [
    ("standby_empty", "Standby 空車"),
    ("standby_full", "Standby 滿料"),
    ("loading", "Processing 裝料"),
    ("out", "Processing 出差中"),
    ("repair", "修車"),
]


STATUS_COLOR = {
    "standby_empty": ft.Colors.GREY_200,
    "standby_full": ft.Colors.GREEN_100,
    "loading": ft.Colors.BLUE_100,
    "out": ft.Colors.ORANGE_100,
    "repair": ft.Colors.RED_100,
}

from config import DATA_DIR, TRIP_EXCEL_PATH, DAILY_EXPORT_DIR

DATA_DIR.mkdir(exist_ok=True)
DAILY_EXPORT_DIR.mkdir(exist_ok=True)


def ensure_trip_excel_exists():
    """
    如果同資料夾沒有「出差資料.xlsx」，自動建立一份。
    """
    if TRIP_EXCEL_PATH.exists():
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "出差資料"

    ws["A1"] = "出差時間"
    ws["B1"] = "車子號碼"
    ws["C1"] = "目的地"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20

    wb.save(TRIP_EXCEL_PATH)




def parse_excel_datetime(value):
    """
    將 Excel A欄的出差時間轉成 datetime。
    支援 Excel 真正時間，也支援文字時間。
    """
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, str):
        value = value.strip()

        formats = [
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
        ]

        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass

    return None


def get_trip_info_from_excel(car_no):
    """
    用車號到「出差資料.xlsx」查：
    A欄：出差時間
    B欄：車子號碼
    C欄：目的地

    如果同車號有多筆，會抓出差時間最新的一筆。
    """
    ensure_trip_excel_exists()

    wb = openpyxl.load_workbook(TRIP_EXCEL_PATH, data_only=True)
    ws = wb.active

    car_no = str(car_no).strip()
    matched_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        trip_time_raw = row[0]
        excel_car_no = row[1]
        destination = row[2]

        if excel_car_no is None:
            continue

        if str(excel_car_no).strip() == car_no:
            trip_time = parse_excel_datetime(trip_time_raw)

            if trip_time is None:
                continue

            matched_list.append({
                "trip_time": trip_time,
                "car_no": str(excel_car_no).strip(),
                "destination": str(destination).strip() if destination else "",
            })

    if not matched_list:
        return None

    latest = max(matched_list, key=lambda x: x["trip_time"])
    return latest

def auto_fill_trip_info(car_no):
    """
    車輛狀態切換成「出差」時：
    1. 用車號到出差資料.xlsx 找資料
    2. 取得出差時間與目的地
    3. 用目的地到 vms.db 查預估分鐘數
    4. 回傳預估回來時間
    """

    trip_info = get_trip_info_from_excel(car_no)

    if trip_info is None:
        return {
            "success": False,
            "message": f"出差資料.xlsx 找不到車號 {car_no} 的出差資料",
            "destination": "",
            "trip_time": None,
            "estimated_return_time": None,
        }

    destination = trip_info["destination"]
    trip_time = trip_info["trip_time"]

    # 新增：檢查「目的地 × 車牌」白名單
    # 沒有在設定頁面勾選的車牌，不允許放入出差中。
    if not is_car_allowed_for_destination(destination, car_no):
        return {
            "success": False,
            "message": f"非白名單：車牌 {car_no} 未被允許前往「{destination}」，不可放入出差中",
            "destination": destination,
            "trip_time": trip_time,
            "estimated_return_time": None,
        }

    # 用 Excel C欄的目的地，到 vms.db 查這個目的地的預估分鐘數
    minutes = get_destination_minutes(destination)

    if minutes is None:
        return {
            "success": False,
            "message": f"目的地「{destination}」尚未設定預估時間",
            "destination": destination,
            "trip_time": trip_time,
            "estimated_return_time": None,
        }

    estimated_return_time = trip_time + timedelta(minutes=minutes)

    return {
        "success": True,
        "message": "",
        "destination": destination,
        "trip_time": trip_time,
        "estimated_return_time": estimated_return_time,
    }

def open_dialog(page, dialog):
    page.overlay.append(dialog)
    dialog.open = True
    page.update()


def close_dialog(page, dialog):
    dialog.open = False
    page.update()


def show_message_dialog(page, title, message):
    dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text(title),
        content=ft.Text(message),
    )

    def close(e):
        close_dialog(page, dialog)

    dialog.actions = [
        ft.ElevatedButton("確定", on_click=close),
    ]

    open_dialog(page, dialog)


def get_plate_no_from_event(page, e):
    plate_no = None

    try:
        if hasattr(e, "data") and e.data:
            plate_no = e.data
    except Exception:
        pass

    try:
        if hasattr(e, "src_id") and e.src_id:
            src_control = page.get_control(e.src_id)
            if hasattr(src_control, "data"):
                plate_no = src_control.data
    except Exception:
        pass

    return plate_no


def get_remaining_text(eta_time):
    if eta_time is None or eta_time == "":
        return "", False

    try:
        eta = datetime.strptime(eta_time, "%Y-%m-%d %H:%M:%S")
        now = datetime.now()
        diff = eta - now
        total_seconds = int(diff.total_seconds())

        if total_seconds < 0:
            overdue_seconds = abs(total_seconds)
            h = overdue_seconds // 3600
            m = (overdue_seconds % 3600) // 60
            return f"超時 {h:02d}:{m:02d}", True

        h = total_seconds // 3600
        m = (total_seconds % 3600) // 60
        return f"剩餘 {h:02d}:{m:02d}", False

    except Exception:
        return "", False

def format_eta_time(eta_time):
    if eta_time is None or eta_time == "":
        return ""

    try:
        eta = datetime.strptime(eta_time, "%Y-%m-%d %H:%M:%S")
        return eta.strftime("%m/%d %H:%M")
    except Exception:
        return ""

def get_selected_plates(page, dragged_plate_no):
    """
    如果拖曳的車有被選取，則回傳所有已選取車牌。
    如果拖曳的車沒有被選取，則只移動該車。
    """
    selected = getattr(page, "selected_plates", set())

    if dragged_plate_no in selected and len(selected) > 0:
        return list(selected)

    return [dragged_plate_no]


def clear_selected_plates(page):
    page.selected_plates = set()


def toggle_vehicle_selection(page, plate_no):
    """
    點擊車卡時，選取 / 取消選取。
    """
    if not hasattr(page, "selected_plates"):
        page.selected_plates = set()

    if plate_no in page.selected_plates:
        page.selected_plates.remove(plate_no)
    else:
        page.selected_plates.add(plate_no)

    refresh_page(page)


def get_vehicle_status_map():
    """
    取得目前所有車輛狀態，供多選拖曳判斷使用。
    """
    rows = get_vehicles()
    return {row[0]: row[2] for row in rows}

def vehicle_card(page, row):
    plate_no = row[0]
    status = row[2]
    eta_time = row[4]
    destination = row[5]
    loading_level = row[6]
    repair_reason = row[7]

    remaining_text, is_overdue = get_remaining_text(eta_time)
    eta_display = format_eta_time(eta_time)

    # 預設樣式
    text_color = ft.Colors.BLACK
    card_bgcolor = ft.Colors.WHITE
    border_color = ft.Colors.GREY_500
    selected_plates = getattr(page, "selected_plates", set())
    is_selected = plate_no in selected_plates

    # 裝料倒數結束：亮綠色
    if status == "loading" and is_overdue:
        card_bgcolor = ft.Colors.LIGHT_GREEN_300
        text_color = ft.Colors.BLACK
        border_color = ft.Colors.GREEN_700

    # 出差中倒數結束：亮紅色，並顯示超時計時
    elif status == "out" and is_overdue:
        card_bgcolor = ft.Colors.RED_400
        text_color = ft.Colors.WHITE
        border_color = ft.Colors.RED_900

    # 其他超時狀態：文字紅色
    elif is_overdue:
        text_color = ft.Colors.RED

    if is_selected:
        border_color = ft.Colors.BLUE_700
        if card_bgcolor == ft.Colors.WHITE:
            card_bgcolor = ft.Colors.BLUE_50

    # 空車 / 滿料：小卡片
    if status in ["standby_empty", "standby_full"]:
        card = ft.Container(
            content=ft.Text(
                plate_no,
                size=11,
                weight=ft.FontWeight.BOLD,
                color=text_color,
            ),
            padding=3,
            margin=2,
            bgcolor=card_bgcolor,
            border_radius=5,
            border=ft.border.all(1, border_color),
            width=75,
            height=45,
            alignment=ft.Alignment(0, 0),
            on_click=lambda e, p=plate_no: toggle_vehicle_selection(page, p),
        )

        return ft.Draggable(
            group="vehicle",
            data=plate_no,
            content=card,
            content_feedback=ft.Container(
                content=ft.Text(
                    plate_no,
                    size=11,
                    weight=ft.FontWeight.BOLD,
                    color=ft.Colors.WHITE,
                ),
                padding=4,
                bgcolor=ft.Colors.BLUE_600,
                border_radius=5,
                width=75,
                height=45,
                alignment=ft.Alignment(0, 0),
            ),
        )

    # 裝料 / 出差中 / 修車：橫向卡片
    row_controls = [
        ft.Text(
            plate_no,
            size=15,
            weight=ft.FontWeight.BOLD,
            color=text_color,
        )
    ]

    if status == "loading":
        if loading_level:
            row_controls.append(
                ft.Text(
                    f"裝料：{loading_level}",
                    size=13,
                    color=text_color,
                )
            )

        if remaining_text:
            if is_overdue:
                display_text = remaining_text.replace("超時", "完成")
            else:
                display_text = remaining_text

            row_controls.append(
                ft.Text(
                    display_text,
                    size=13,
                    weight=ft.FontWeight.BOLD if is_overdue else None,
                    color=text_color,
                )
            )

    elif status == "out":
        if destination:
            row_controls.append(
                ft.Text(
                    f"目的地：{destination}",
                    size=13,
                    color=text_color,
                )
            )

        if remaining_text:
            row_controls.append(
                ft.Text(
                    remaining_text,
                    size=13,
                    weight=ft.FontWeight.BOLD if is_overdue else None,
                    color=text_color,
                )
            )

        if eta_display:
            row_controls.append(
                ft.Text(
                    f"預計抵達：{eta_display}",
                    size=13,
                    color=text_color,
                )
            )

    elif status == "repair":
        if repair_reason:
            row_controls.append(
                ft.Text(
                    f"修車：{repair_reason}",
                    size=13,
                    color=text_color,
                )
            )

    else:
        if remaining_text:
            row_controls.append(
                ft.Text(
                    remaining_text,
                    size=13,
                    color=text_color,
                )
            )

    # 這裡一定要放在 if / elif / else 外面
    # 確保 loading / out / repair 都一定會建立 card
    if status == "loading":
        card_width = 430
    elif status == "out":
        card_width = 620
    elif status == "repair":
        card_width = 520
    else:
        card_width = 430

    card = ft.Container(
        content=ft.Row(
            controls=row_controls,
            spacing=12,
            alignment=ft.MainAxisAlignment.START,
        ),
        padding=8,
        margin=4,
        bgcolor=card_bgcolor,
        border_radius=8,
        border=ft.border.all(1, border_color),
        width=card_width,
        height=48,
        alignment=ft.Alignment(-1, 0),
        on_click=lambda e, p=plate_no: toggle_vehicle_selection(page, p),
    )

    return ft.Draggable(
        group="vehicle",
        data=plate_no,
        content=card,
        content_feedback=ft.Container(
            content=ft.Text(
                plate_no,
                size=15,
                weight=ft.FontWeight.BOLD,
                color=ft.Colors.WHITE,
            ),
            padding=8,
            bgcolor=ft.Colors.BLUE_600,
            border_radius=8,
            width=340,
            height=48,
            alignment=ft.Alignment(0, 0),
        ),
    )

def show_out_dialog(page, plate_no):
    plate_nos = plate_no if isinstance(plate_no, list) else [plate_no]

    success_count = 0
    error_messages = []

    for p in plate_nos:
        trip = auto_fill_trip_info(p)

        if not trip["success"]:
            error_messages.append(f"{p}：{trip['message']}")
            continue

        destination = trip["destination"]
        trip_time = trip["trip_time"]
        estimated_return_time = trip["estimated_return_time"]

        remaining_hours = (estimated_return_time - datetime.now()).total_seconds() / 3600

        try:
            trip_time_text = trip_time.strftime("%Y-%m-%d %H:%M:%S")
            update_vehicle_out(p, destination, remaining_hours, trip_time_text)
            success_count += 1

        except Exception as ex:
            error_messages.append(f"{p}：{str(ex)}")

    clear_selected_plates(page)
    refresh_page(page)

    if error_messages:
        show_message_dialog(
            page,
            "部分車輛出差設定失敗",
            "\n".join(error_messages)
        )

def show_loading_dialog(page, plate_no):
    plate_nos = plate_no if isinstance(plate_no, list) else [plate_no]

    loading_levels = get_loading_levels()
    loading_label_map = {}

    for level, hours in loading_levels:
        label = f"{level}（{hours} hr）"
        loading_label_map[label] = {
            "level": level,
            "hours": hours,
        }

    labels = list(loading_label_map.keys())

    if not labels:
        show_message_dialog(page, "無裝料等級設定", "請先到設定頁面新增裝料等級。")
        return

    dropdown = ft.Dropdown(
        label="裝料等級",
        options=[ft.dropdown.Option(label) for label in labels],
        value=labels[0],
        width=330,
    )

    title_text = f"{plate_nos[0]} 裝料設定" if len(plate_nos) == 1 else f"{len(plate_nos)} 台車裝料設定"

    dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text(title_text),
        content=ft.Column(
            controls=[
                ft.Text(
                    "選取車輛：" + "、".join(plate_nos),
                    size=12,
                    color=ft.Colors.GREY_700,
                ),
                dropdown,
                ft.Text(
                    "時間會直接套用設定頁面的「裝料等級對應小時」。",
                    size=12,
                    color=ft.Colors.GREY_600,
                ),
            ],
            tight=True,
        ),
    )

    def confirm(e):
        selected_label = dropdown.value
        selected_data = loading_label_map[selected_label]

        level = selected_data["level"]
        loading_hours = selected_data["hours"]
        loading_minutes = float(loading_hours) * 60

        close_dialog(page, dialog)

        for p in plate_nos:
            update_vehicle_loading(p, level, loading_minutes)

        clear_selected_plates(page)
        refresh_page(page)

    def cancel(e):
        close_dialog(page, dialog)

    dialog.actions = [
        ft.TextButton("取消", on_click=cancel),
        ft.ElevatedButton("確定", on_click=confirm),
    ]

    open_dialog(page, dialog)




def parse_actual_finish_time(value):
    value = str(value).strip()

    formats = [
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass

    return None

def show_complete_out_dialog(page, plate_no, target_status):
    plate_nos = plate_no if isinstance(plate_no, list) else [plate_no]

    now_text = datetime.now().strftime("%Y/%m/%d %H:%M")

    actual_time_text = ft.TextField(
        label="實際完成時間",
        value=now_text,
        width=300,
        hint_text="例如：2026/04/27 15:30",
    )

    title_text = f"{plate_nos[0]} 出差完成紀錄" if len(plate_nos) == 1 else f"{len(plate_nos)} 台車出差完成紀錄"

    dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text(title_text),
        content=ft.Column(
            controls=[
                ft.Text(
                    "選取車輛：" + "、".join(plate_nos),
                    size=12,
                    color=ft.Colors.GREY_700,
                ),
                ft.Text(
                    "請輸入實際完成時間。確認後，車輛才會移出「出差中」。",
                    size=13,
                    color=ft.Colors.GREY_700,
                ),
                actual_time_text,
            ],
            tight=True,
            spacing=10,
        ),
    )

    def confirm(e):
        actual_dt = parse_actual_finish_time(actual_time_text.value)

        if actual_dt is None:
            actual_time_text.error_text = "時間格式錯誤，請輸入例如：2026/04/27 15:30"
            page.update()
            return

        actual_finish_time = actual_dt.strftime("%Y-%m-%d %H:%M:%S")

        try:
            close_dialog(page, dialog)

            for p in plate_nos:
                complete_vehicle_out(p, target_status, actual_finish_time)

            clear_selected_plates(page)
            refresh_page(page)

        except Exception as ex:
            show_message_dialog(page, "出差完成紀錄失敗", str(ex))

    def cancel(e):
        close_dialog(page, dialog)

    dialog.actions = [
        ft.TextButton("取消", on_click=cancel),
        ft.ElevatedButton("確定", on_click=confirm),
    ]

    open_dialog(page, dialog)


def show_repair_dialog(page, plate_no):
    plate_nos = plate_no if isinstance(plate_no, list) else [plate_no]

    reason_text = ft.TextField(
        label="修車原因",
        hint_text="例如：輪胎異常、煞車異常、引擎問題",
        multiline=True,
        width=350,
    )

    title_text = f"{plate_nos[0]} 修車原因" if len(plate_nos) == 1 else f"{len(plate_nos)} 台車修車原因"

    dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text(title_text),
        content=ft.Column(
            controls=[
                ft.Text(
                    "選取車輛：" + "、".join(plate_nos),
                    size=12,
                    color=ft.Colors.GREY_700,
                ),
                reason_text,
            ],
            tight=True,
        ),
    )

    def confirm(e):
        reason = reason_text.value.strip()

        if reason == "":
            reason_text.error_text = "修車原因必填"
            page.update()
            return

        close_dialog(page, dialog)

        for p in plate_nos:
            update_vehicle_repair(p, reason)

        clear_selected_plates(page)
        refresh_page(page)

    def cancel(e):
        close_dialog(page, dialog)

    dialog.actions = [
        ft.TextButton("取消", on_click=cancel),
        ft.ElevatedButton("確定", on_click=confirm),
    ]

    open_dialog(page, dialog)


def status_column(page, status_key, title, vehicles):
    def on_accept(e):
        if not can_edit(page):
            show_message_dialog(page, "權限不足", "你的權限為瀏覽者，只能查看狀態，不能拖曳修改。")
            return

        plate_no = get_plate_no_from_event(page, e)

        if plate_no is None or str(plate_no).strip() == "":
            print("拖拉失敗：沒有取得車牌號碼")
            return

        target_plates = get_selected_plates(page, plate_no)

        status_map = get_vehicle_status_map()

        # 過濾不存在的車牌
        target_plates = [p for p in target_plates if p in status_map]

        if not target_plates:
            return

        current_status_set = set(status_map[p] for p in target_plates)

        # 為了避免混亂，多選時只允許相同目前狀態的車一起拖
        if len(current_status_set) > 1:
            show_message_dialog(
                page,
                "多選狀態不同",
                "多選拖曳時，請只選取目前在同一個狀態區塊的車輛。"
            )
            return

        current_status = list(current_status_set)[0]

        if current_status == status_key:
            print("選取車輛已經在目標狀態，不重新設定")
            return

        # 如果目前在出差中，要移出時必須填實際完成時間
        if current_status == "out" and status_key != "out":
            show_complete_out_dialog(page, target_plates, status_key)
            return

        if status_key == "out":
            show_out_dialog(page, target_plates)

        elif status_key == "loading":
            show_loading_dialog(page, target_plates)

        elif status_key == "repair":
            show_repair_dialog(page, target_plates)

        else:
            for p in target_plates:
                update_vehicle_status(p, status_key)

            clear_selected_plates(page)
            refresh_page(page)

    vehicle_controls = []

    sorted_vehicles = sorted(
        vehicles,
        key=lambda r: (
            0 if get_remaining_text(r[4])[1] else 1,
            r[4] if r[4] else "9999-12-31 23:59:59",
            r[3] if r[3] else "9999-12-31 23:59:59",
        ),
    )

    for row in sorted_vehicles:
        status = row[2]

        if status == status_key:
            vehicle_controls.append(vehicle_card(page, row))

    # 空車 / 滿料：小卡片，每列兩台
    if status_key in ["standby_empty", "standby_full"]:
        vehicle_area = ft.GridView(
            controls=vehicle_controls,
            runs_count=2,
            max_extent=90,
            child_aspect_ratio=1.65,
            spacing=4,
            run_spacing=4,
            expand=True,
        )
        column_width = 200

    else:
        # 裝料 / 出差中 / 修車：
        # 外層 Row 負責左右捲動
        # 內層 Column 負責上下排列車卡
        if status_key == "loading":
            column_width = 340
            vehicle_list_width = 450
        elif status_key == "out":
            column_width = 520
            vehicle_list_width = 650
        elif status_key == "repair":
            column_width = 400
            vehicle_list_width = 550
        else:
            column_width = 380
            vehicle_list_width = 450

        vehicle_list = ft.Column(
            controls=vehicle_controls,
            scroll=ft.ScrollMode.AUTO,
            spacing=4,
            width=vehicle_list_width,
            
        )

        vehicle_area = ft.Row(
            controls=[vehicle_list],
            scroll=ft.ScrollMode.AUTO,
            expand=True,
            vertical_alignment=ft.CrossAxisAlignment.START,
        )

    column_content = ft.Container(
        content=ft.Column(
            controls=[
                ft.Text(title, size=15, weight=ft.FontWeight.BOLD),
                ft.Divider(),
                vehicle_area,
            ],
            spacing=5,
        ),
        width=column_width,
        height=430,
        padding=10,
        bgcolor=STATUS_COLOR.get(status_key, ft.Colors.GREY_100),
        border_radius=10,
        border=ft.border.all(1, ft.Colors.GREY_400),
    )

    return ft.DragTarget(
        group="vehicle",
        content=column_content,
        on_accept=on_accept,
    )


def build_material_area(page, material_type, all_vehicles):
    material_vehicles = []

    for row in all_vehicles:
        if row[1] == material_type:
            material_vehicles.append(row)

    controls = []

    for status_key, status_name in STATUS_LIST:
        controls.append(
            status_column(
                page=page,
                status_key=status_key,
                title=status_name,
                vehicles=material_vehicles,
            )
        )

    return ft.Column(
        controls=[
            ft.Text(f"{material_type} 料件區", size=22, weight=ft.FontWeight.BOLD),
            ft.Row(controls=controls, scroll=ft.ScrollMode.AUTO),
        ],
        spacing=10,
    )


def build_vehicle_setting_section(page):
    vehicles = get_vehicles()
    rows = []
    edit_controls = []

    for row in vehicles:
        old_plate_no = row[0]
        old_material_type = row[1]

        plate_text = ft.TextField(
            value=old_plate_no,
            width=130,
            dense=True,
            text_size=13,
        )

        material_dropdown = ft.Dropdown(
            options=[
                ft.dropdown.Option("A"),
                ft.dropdown.Option("B"),
            ],
            value=old_material_type,
            width=90,
            dense=True,
        )

        edit_controls.append({
            "old_plate_no": old_plate_no,
            "plate_text": plate_text,
            "material_dropdown": material_dropdown,
        })

        def make_delete_func(plate):
            def do_delete(e):
                try:
                    delete_vehicle(plate)
                    show_settings_page(page)
                except Exception as ex:
                    show_message_dialog(page, "刪除失敗", str(ex))

            return do_delete

        rows.append(
            ft.Row(
                controls=[
                    ft.Text("車牌", size=13),
                    plate_text,
                    ft.Text("料件", size=13),
                    material_dropdown,
                    ft.TextButton("刪除", on_click=make_delete_func(old_plate_no)),
                ],
                spacing=8,
                wrap=True,
            )
        )

    new_plate_text = ft.TextField(
        label="新增車牌",
        width=140,
        dense=True,
    )

    new_material_dropdown = ft.Dropdown(
        label="料件",
        options=[
            ft.dropdown.Option("A"),
            ft.dropdown.Option("B"),
        ],
        value="A",
        width=100,
        dense=True,
    )

    def add_new_vehicle(e):
        plate = new_plate_text.value.strip()
        material = new_material_dropdown.value

        if plate == "":
            show_message_dialog(page, "錯誤", "新增車牌不可空白。")
            return

        try:
            add_vehicle(plate, material)
            show_settings_page(page)
        except Exception as ex:
            show_message_dialog(page, "新增失敗", str(ex))

    def save_vehicle_settings(e):
        try:
            for item in edit_controls:
                old_plate = item["old_plate_no"]
                new_plate = item["plate_text"].value.strip()
                new_material = item["material_dropdown"].value

                if new_plate == "":
                    show_message_dialog(page, "錯誤", "車牌不可空白。")
                    return

                update_vehicle(old_plate, new_plate, new_material)

            show_message_dialog(page, "儲存完成", "車輛設定已儲存。")
            show_settings_page(page)

        except Exception as ex:
            show_message_dialog(page, "儲存失敗", str(ex))

    return ft.Container(
        content=ft.Column(
            controls=[
                ft.Text("車輛設定", size=18, weight=ft.FontWeight.BOLD),
                ft.Row(
                    controls=[
                        new_plate_text,
                        new_material_dropdown,
                        ft.ElevatedButton("新增車輛", on_click=add_new_vehicle),
                    ],
                    spacing=8,
                    wrap=True,
                ),
                ft.Divider(),
                *rows,
                ft.Row(
                    controls=[
                        ft.ElevatedButton("儲存車輛設定", on_click=save_vehicle_settings),
                    ],
                    alignment=ft.MainAxisAlignment.END,
                ),
            ],
            spacing=8,
        ),
        padding=12,
        bgcolor=ft.Colors.GREY_100,
        border_radius=10,
        border=ft.border.all(1, ft.Colors.GREY_400),
    )


def build_destination_setting_section(page):
    destinations = get_destinations()
    rows = []
    edit_controls = []

    for name, hours in destinations:
        old_name = name

        name_text = ft.TextField(
            value=name,
            width=140,
            dense=True,
            text_size=13,
        )

        hours_text = ft.TextField(
            value=str(hours),
            width=100,
            dense=True,
            text_size=13,
        )

        edit_controls.append({
            "old_name": old_name,
            "name_text": name_text,
            "hours_text": hours_text,
        })

        def make_delete_func(dest):
            def do_delete(e):
                try:
                    delete_destination(dest)
                    show_settings_page(page)
                except Exception as ex:
                    show_message_dialog(page, "刪除失敗", str(ex))

            return do_delete

        rows.append(
            ft.Container(
                content=ft.Column(
                    controls=[
                        ft.Row(
                            controls=[
                                ft.Text("目的地", size=13),
                                name_text,
                                ft.Text("時數 hr", size=13),
                                hours_text,
                                ft.TextButton("刪除", on_click=make_delete_func(old_name)),
                            ],
                            spacing=8,
                            wrap=True,
                        ),
                        build_destination_whitelist_area(page, old_name),
                    ],
                    spacing=8,
                ),
                padding=8,
                bgcolor=ft.Colors.WHITE,
                border_radius=8,
                border=ft.border.all(1, ft.Colors.GREY_300),
            )
        )

    new_name_text = ft.TextField(
        label="新增目的地",
        width=150,
        dense=True,
    )

    new_hours_text = ft.TextField(
        label="時數 hr",
        value="1",
        width=100,
        dense=True,
    )

    def add_new_destination(e):
        name = new_name_text.value.strip()

        if name == "":
            show_message_dialog(page, "錯誤", "目的地名稱不可空白。")
            return

        try:
            add_destination(name, float(new_hours_text.value))
            show_settings_page(page)
        except Exception as ex:
            show_message_dialog(page, "新增失敗", str(ex))

    def save_destination_settings(e):
        try:
            for item in edit_controls:
                old_name = item["old_name"]
                new_name = item["name_text"].value.strip()
                hours_value = item["hours_text"].value

                if new_name == "":
                    show_message_dialog(page, "錯誤", "目的地名稱不可空白。")
                    return

                update_destination(old_name, new_name, float(hours_value))

            show_message_dialog(page, "儲存完成", "出差目的地設定已儲存。")
            show_settings_page(page)

        except Exception as ex:
            show_message_dialog(page, "儲存失敗", str(ex))

    return ft.Container(
        content=ft.Column(
            controls=[
                ft.Text("出差目的地設定", size=18, weight=ft.FontWeight.BOLD),
                ft.Row(
                    controls=[
                        new_name_text,
                        new_hours_text,
                        ft.ElevatedButton("新增目的地", on_click=add_new_destination),
                    ],
                    spacing=8,
                    wrap=True,
                ),
                ft.Divider(),
                *rows,
                ft.Row(
                    controls=[
                        ft.ElevatedButton("儲存目的地設定", on_click=save_destination_settings),
                    ],
                    alignment=ft.MainAxisAlignment.END,
                ),
            ],
            spacing=8,
        ),
        padding=12,
        bgcolor=ft.Colors.GREY_100,
        border_radius=10,
        border=ft.border.all(1, ft.Colors.GREY_400),
    )

def build_user_setting_section(page):
    if not is_admin(page):
        return ft.Container()

    users = get_users()
    rows = []
    edit_controls = []

    for login_account, display_name, role, is_active in users:
        old_login = login_account

        account_text = ft.TextField(
            value=login_account,
            width=130,
            dense=True,
            text_size=13,
        )

        name_text = ft.TextField(
            value=display_name,
            width=120,
            dense=True,
            text_size=13,
        )

        password_text = ft.TextField(
            label="新密碼",
            password=True,
            can_reveal_password=True,
            width=160,
            dense=True,
            text_size=13,
        )

        role_dropdown = ft.Dropdown(
            options=[
                ft.dropdown.Option("admin"),
                ft.dropdown.Option("editor"),
                ft.dropdown.Option("viewer"),
            ],
            value=role,
            width=130,
            dense=True,
        )

        active_dropdown = ft.Dropdown(
            options=[
                ft.dropdown.Option("1"),
                ft.dropdown.Option("0"),
            ],
            value=str(is_active),
            width=80,
            dense=True,
        )

        edit_controls.append({
            "old_login": old_login,
            "account_text": account_text,
            "name_text": name_text,
            "password_text": password_text,
            "role_dropdown": role_dropdown,
            "active_dropdown": active_dropdown,
        })

        def make_delete_func(account):
            def do_delete(e):
                if account == page.current_user["login_account"]:
                    show_message_dialog(page, "不可刪除", "不可刪除目前登入中的帳號。")
                    return

                try:
                    delete_user(account)
                    show_settings_page(page)
                except Exception as ex:
                    show_message_dialog(page, "刪除失敗", str(ex))

            return do_delete

        rows.append(
            ft.Row(
                controls=[
                    ft.Text("帳號", size=13),
                    account_text,
                    ft.Text("姓名", size=13),
                    name_text,
                    password_text,
                    ft.Text("權限", size=13),
                    role_dropdown,
                    ft.Text("啟用", size=13),
                    active_dropdown,
                    ft.TextButton("刪除", on_click=make_delete_func(old_login)),
                ],
                spacing=8,
                wrap=True,
            )
        )

    new_account_text = ft.TextField(
        label="新增帳號",
        width=130,
        dense=True,
    )

    new_name_text = ft.TextField(
        label="姓名",
        width=150,
        dense=True,
    )

    new_password_text = ft.TextField(
        label="密碼",
        password=True,
        can_reveal_password=True,
        width=120,
        dense=True,
    )

    new_role_dropdown = ft.Dropdown(
        label="權限",
        options=[
            ft.dropdown.Option("admin"),
            ft.dropdown.Option("editor"),
            ft.dropdown.Option("viewer"),
        ],
        value="viewer",
        width=120,
        dense=True,
    )

    def add_new_user(e):
        account = new_account_text.value.strip()
        name = new_name_text.value.strip()
        password = new_password_text.value.strip()
        role = new_role_dropdown.value

        if account == "" or name == "" or password == "":
            show_message_dialog(page, "錯誤", "帳號、姓名、密碼都必填。")
            return

        try:
            add_user(account, name, password, role)
            show_settings_page(page)
        except Exception as ex:
            show_message_dialog(page, "新增失敗", str(ex))

    def save_user_settings(e):
        try:
            for item in edit_controls:
                old_login = item["old_login"]
                new_login = item["account_text"].value.strip()
                display_name = item["name_text"].value.strip()
                password = item["password_text"].value.strip()
                role = item["role_dropdown"].value
                is_active = item["active_dropdown"].value

                if new_login == "" or display_name == "":
                    show_message_dialog(page, "錯誤", "帳號與姓名不可空白。")
                    return

                update_user(
                    old_login,
                    new_login,
                    display_name,
                    password,
                    role,
                    is_active,
                )

            show_message_dialog(page, "儲存完成", "使用者設定已儲存。")
            show_settings_page(page)

        except Exception as ex:
            show_message_dialog(page, "儲存失敗", str(ex))

    return ft.Container(
        content=ft.Column(
            controls=[
                ft.Text("使用者管理", size=18, weight=ft.FontWeight.BOLD),
                ft.Row(
                    controls=[
                        new_account_text,
                        new_name_text,
                        new_password_text,
                        new_role_dropdown,
                        ft.ElevatedButton("新增使用者", on_click=add_new_user),
                    ],
                    spacing=8,
                    wrap=True,
                ),
                ft.Divider(),
                *rows,
                ft.Row(
                    controls=[
                        ft.ElevatedButton("儲存使用者設定", on_click=save_user_settings),
                    ],
                    alignment=ft.MainAxisAlignment.END,
                ),
            ],
            spacing=8,
        ),
        padding=12,
        bgcolor=ft.Colors.GREY_100,
        border_radius=10,
        border=ft.border.all(1, ft.Colors.GREY_400),
    )

def build_loading_setting_section(page):
    loading_levels = get_loading_levels()
    rows = []
    edit_controls = []

    for level_name, hours in loading_levels:
        old_level = level_name

        level_text = ft.TextField(
            value=level_name,
            width=140,
            dense=True,
            text_size=13,
        )

        hours_text = ft.TextField(
            value=str(hours),
            width=100,
            dense=True,
            text_size=13,
        )

        edit_controls.append({
            "old_level": old_level,
            "level_text": level_text,
            "hours_text": hours_text,
        })

        def make_delete_func(level):
            def do_delete(e):
                try:
                    delete_loading_level(level)
                    show_settings_page(page)
                except Exception as ex:
                    show_message_dialog(page, "刪除失敗", str(ex))

            return do_delete

        rows.append(
            ft.Row(
                controls=[
                    ft.Text("等級", size=13),
                    level_text,
                    ft.Text("小時 hr", size=13),
                    hours_text,
                    ft.TextButton("刪除", on_click=make_delete_func(old_level)),
                ],
                spacing=8,
                wrap=True,
            )
        )

    new_level_text = ft.TextField(
        label="新增裝料等級",
        width=150,
        dense=True,
    )

    new_hours_text = ft.TextField(
        label="小時 hr",
        value="0.5",
        width=100,
        dense=True,
    )

    def add_new_loading_level(e):
        level = new_level_text.value.strip()

        if level == "":
            show_message_dialog(page, "錯誤", "裝料等級不可空白。")
            return

        try:
            add_loading_level(level, float(new_hours_text.value))
            show_settings_page(page)
        except Exception as ex:
            show_message_dialog(page, "新增失敗", str(ex))

    def save_loading_settings(e):
        try:
            for item in edit_controls:
                old_level = item["old_level"]
                new_level = item["level_text"].value.strip()
                hours_value = item["hours_text"].value

                if new_level == "":
                    show_message_dialog(page, "錯誤", "裝料等級名稱不可空白。")
                    return

                update_loading_level(old_level, new_level, float(hours_value))

            show_message_dialog(page, "儲存完成", "裝料等級設定已儲存。")
            show_settings_page(page)

        except Exception as ex:
            show_message_dialog(page, "儲存失敗", str(ex))

    return ft.Container(
        content=ft.Column(
            controls=[
                ft.Text("裝料等級設定", size=18, weight=ft.FontWeight.BOLD),
                ft.Row(
                    controls=[
                        new_level_text,
                        new_hours_text,
                        ft.ElevatedButton("新增裝料等級", on_click=add_new_loading_level),
                    ],
                    spacing=8,
                    wrap=True,
                ),
                ft.Divider(),
                *rows,
                ft.Row(
                    controls=[
                        ft.ElevatedButton("儲存裝料等級設定", on_click=save_loading_settings),
                    ],
                    alignment=ft.MainAxisAlignment.END,
                ),
            ],
            spacing=8,
        ),
        padding=12,
        bgcolor=ft.Colors.GREY_100,
        border_radius=10,
        border=ft.border.all(1, ft.Colors.GREY_400),
    )

def can_edit(page):
    if not hasattr(page, "current_user") or page.current_user is None:
        return False

    return page.current_user["role"] in ["admin", "editor"]


def is_admin(page):
    if not hasattr(page, "current_user") or page.current_user is None:
        return False

    return page.current_user["role"] == "admin"

def show_login_page(page):
    page.controls.clear()
    page.current_view = "login"

    account_text = ft.TextField(
        label="登入帳號",
        width=280,
    )

    password_text = ft.TextField(
        label="登入密碼",
        password=True,
        can_reveal_password=True,
        width=280,
    )

    message_text = ft.Text(
        "",
        color=ft.Colors.RED,
        size=13,
    )

    def do_login(e):
        account = account_text.value.strip()
        password = password_text.value.strip()

        user = authenticate_user(account, password)

        if user is None:
            message_text.value = "帳號或密碼錯誤，或帳號已停用。"
            page.update()
            return

        page.current_user = user
        show_main_page(page)

    login_box = ft.Container(
        content=ft.Column(
            controls=[
                ft.Text("車輛管理系統 VMS", size=28, weight=ft.FontWeight.BOLD),
                ft.Text("請登入", size=18),
                account_text,
                password_text,
                ft.ElevatedButton("登入", on_click=do_login),
                message_text,
                ft.Text(
                    "初始管理員：admin / 1234。首次登入後請到使用者管理修改密碼。",
                    size=12,
                    color=ft.Colors.GREY_700,
                ),
            ],
            spacing=12,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        padding=30,
        width=420,
        bgcolor=ft.Colors.GREY_100,
        border_radius=12,
        border=ft.border.all(1, ft.Colors.GREY_400),
    )

    page.add(
        ft.Row(
            controls=[login_box],
            alignment=ft.MainAxisAlignment.CENTER,
        )
    )

    page.update()

def logout(page):
    page.current_user = None
    show_login_page(page)

def show_main_page(page):
    page.controls.clear()
    page.current_view = "main"

    all_vehicles = get_vehicles()

    user_text = "未登入"

    if hasattr(page, "current_user") and page.current_user is not None:
        user_text = f"登入者：{page.current_user['display_name']} / 權限：{page.current_user['role']}"

    selected_count = len(getattr(page, "selected_plates", set()))

    buttons = [
        ft.Text(
            user_text,
            size=13,
            color=ft.Colors.GREY_700,
        ),
        ft.Text(
            f"已選取：{selected_count} 台",
            size=13,
            color=ft.Colors.BLUE_700,
        ),
        ft.ElevatedButton(
            "清除選取",
            on_click=lambda e: (clear_selected_plates(page), refresh_page(page)),
        ),
        ft.ElevatedButton(
            "重新整理",
            on_click=lambda e: show_main_page(page),
        ),
        ft.ElevatedButton(
            "匯出出差紀錄",
            on_click=lambda e: show_export_trip_records_dialog(page),
        ),
        ft.ElevatedButton(
            "登出",
            on_click=lambda e: logout(page),
        ),
    ]

    if can_edit(page):
        buttons.insert(
            1,
            ft.ElevatedButton(
                "設定頁面",
                on_click=lambda e: show_settings_page(page),
            )
        )



    content = ft.Column(
        controls=[
            ft.Text("車輛管理系統 VMS", size=28, weight=ft.FontWeight.BOLD),
            ft.Row(
                controls=buttons,
                spacing=10,
                wrap=True,
            ),
            ft.Divider(),
            build_material_area(page, "A", all_vehicles),
            ft.Divider(),
            build_material_area(page, "B", all_vehicles),
        ],
        spacing=10,
        scroll=ft.ScrollMode.AUTO,
    )

    page.main_screenshot = ft.Screenshot(content=content)

    page.add(page.main_screenshot)
    page.update()


def show_settings_page(page):
    page.controls.clear()
    page.current_view = "settings"

    user_text = "未登入"
    if hasattr(page, "current_user") and page.current_user is not None:
        user_text = f"登入者：{page.current_user['display_name']} / 權限：{page.current_user['role']}"

    content = ft.Column(
        controls=[
            ft.Text("設定頁面", size=28, weight=ft.FontWeight.BOLD),
            ft.Row(
                controls=[
                    ft.ElevatedButton(
                        "返回主畫面",
                        on_click=lambda e: show_main_page(page),
                    ),
                    ft.ElevatedButton(
                        "重新整理設定",
                        on_click=lambda e: show_settings_page(page),
                    ),
                    ft.ElevatedButton(
                        "登出",
                        on_click=lambda e: logout(page),
                    ),
                ],
                spacing=10,
                wrap=True,
            ),
            ft.Text(
                user_text,
                size=13,
                color=ft.Colors.GREY_700,
            ),
            ft.Text(
                "在這裡設定車輛、出差目的地與裝料等級。主畫面的 Processing 選單會直接讀取這裡的設定。",
                size=13,
                color=ft.Colors.GREY_700,
            ),
            ft.Divider(),
            ft.Row(
                controls=[
                    ft.Container(
                        content=build_vehicle_setting_section(page),
                        expand=1,
                    ),
                    ft.Container(
                        content=build_destination_setting_section(page),
                        expand=1,
                    ),
                    ft.Container(
                        content=build_loading_setting_section(page),
                        expand=1,
                    ),
                ],
                spacing=15,
                vertical_alignment=ft.CrossAxisAlignment.START,
            ),
            ft.Divider(),
            build_user_setting_section(page) if is_admin(page) else ft.Container(),
        ],
        spacing=10,
        scroll=ft.ScrollMode.AUTO,
    )

    page.add(content)
    page.update()
    

def refresh_page(page):
    if getattr(page, "current_view", "main") == "settings":
        show_settings_page(page)
    else:
        show_main_page(page)




def start_auto_refresh(page):
    def auto_refresh_loop():
        while True:
            time.sleep(1800)

            try:
                # 只在主畫面自動刷新，避免你在設定頁面打字時被刷新掉
                if getattr(page, "current_view", "main") == "main":
                    show_main_page(page)
            except Exception as ex:
                print("自動刷新錯誤：", ex)

    t = threading.Thread(target=auto_refresh_loop, daemon=True)
    t.start()

async def capture_main_status_screen(page):
    """
    截圖目前 VMS 主狀態畫面，輸出 PNG 到：
    main.py 同資料夾 / 每日自動匯出
    """

    try:
        # 不截設定頁面。
        # 如果當下在設定頁面，先切回主狀態畫面再截圖。
        if getattr(page, "current_view", "main") != "main":
            show_main_page(page)
            await asyncio.sleep(0.8)

        screenshot_control = getattr(page, "main_screenshot", None)

        if screenshot_control is None:
            print("每日自動截圖失敗：找不到 main_screenshot")
            return

        DAILY_EXPORT_DIR.mkdir(exist_ok=True)

        file_name = f"VMS狀態畫面_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        export_path = DAILY_EXPORT_DIR / file_name

        png_bytes = await screenshot_control.capture(pixel_ratio=2)

        with open(export_path, "wb") as f:
            f.write(png_bytes)

        print(f"每日自動截圖完成：{export_path}")

    except Exception as ex:
        print("每日自動截圖錯誤：", ex)


def start_daily_auto_export(page):
    """
    每天 23:59 自動截圖一次。
    使用背景 thread 檢查時間，再用 page.run_task 執行 Flet 的 async 截圖。
    """

    def auto_export_loop():
        last_export_date = None

        while True:
            try:
                now = datetime.now()

                # 每天 23:59:00 ~ 23:59:59 之間觸發一次
                if now.hour == 23 and now.minute == 59:
                    if last_export_date != now.date():
                        page.run_task(capture_main_status_screen, page)
                        last_export_date = now.date()

                time.sleep(20)

            except Exception as ex:
                print("每日自動匯出排程錯誤：", ex)
                time.sleep(60)

    t = threading.Thread(target=auto_export_loop, daemon=True)
    t.start()

def main(page: ft.Page):
    page.title = "車輛管理系統 VMS"
    page.window_width = 1250
    page.window_height = 950
    page.padding = 20
    page.scroll = ft.ScrollMode.AUTO

    import database
    print("main.py 使用的 database.py 位置：", database.__file__)
    print("main.py 使用的 vms.db 位置：", database.DB_NAME)

    init_db()
    seed_default_admin()

    page.current_user = None
    page.current_view = "login"
    page.selected_plates = set()

    show_login_page(page)

    start_auto_refresh(page)
    start_daily_auto_export(page)
def build_destination_whitelist_area(page, destination_name):
    """
    建立某目的地底下的車牌白名單 Checkbox 區塊。
    勾選後自動儲存，車牌橫向排列。
    """
    all_cars = get_all_enabled_car_nos()
    selected_cars = get_destination_whitelist(destination_name)

    checkbox_map = {}
    checkbox_controls = []

    def auto_save_whitelist(e=None):
        selected = []

        for car_no, cb in checkbox_map.items():
            if cb.value:
                selected.append(car_no)

        set_destination_whitelist(destination_name, selected)

        page.snack_bar = ft.SnackBar(
            ft.Text(f"已自動儲存「{destination_name}」的車牌白名單")
        )
        page.snack_bar.open = True
        page.update()

    for car_no in all_cars:
        cb = ft.Checkbox(
            value=str(car_no) in selected_cars,
            on_change=auto_save_whitelist,
            width=35,
        )

        checkbox_map[str(car_no)] = cb

        checkbox_controls.append(
            ft.Container(
                content=ft.Row(
                    controls=[
                        cb,
                        ft.Text(
                            str(car_no),
                            size=12,  # 車牌字體大小在這裡改
                        ),
                    ],
                    spacing=0,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
                width=90,
                padding=0,
            )
        )

    if not checkbox_controls:
        checkbox_controls.append(
            ft.Text(
                "目前尚無啟用車牌；請先在左側車輛設定新增車牌。",
                size=12,
                color=ft.Colors.RED_400,
            )
        )

    return ft.Container(
        content=ft.Column(
            controls=[
                ft.Text(
                    "允許出差車牌",
                    size=14,
                    weight=ft.FontWeight.BOLD,
                ),
                ft.Text(
                    "勾選後會自動儲存；只有勾選的車牌可以放入此目的地的出差中。",
                    size=12,
                    color=ft.Colors.GREY_600,
                ),
                ft.Row(
                    controls=checkbox_controls,
                    wrap=True,
                    spacing=4,
                    run_spacing=0,
                ),
            ],
            spacing=6,
        ),
        padding=8,
        border=ft.border.all(1, ft.Colors.GREY_300),
        border_radius=8,
    )







def parse_export_datetime(value, is_end=False):
    """
    匯出時間區間解析。
    支援：
    2026/04/27
    2026/04/27 08:00
    2026/04/27 08:00:00
    2026-04-27
    2026-04-27 08:00
    2026-04-27 08:00:00

    如果只輸入日期：
    起始時間 → 00:00:00
    結束時間 → 23:59:59
    """
    value = str(value).strip()

    if value == "":
        return None

    date_formats = [
        "%Y/%m/%d",
        "%Y-%m-%d",
    ]

    datetime_formats = [
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    ]

    for fmt in datetime_formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass

    for fmt in date_formats:
        try:
            dt = datetime.strptime(value, fmt)

            if is_end:
                return dt.replace(hour=23, minute=59, second=59)
            else:
                return dt.replace(hour=0, minute=0, second=0)

        except ValueError:
            pass

    return None


def format_duration_text(start_time_text, finish_time_text):
    """
    回傳 hh:mm 文字，主要給無法轉成 Excel 時間時備用。
    """
    try:
        start_dt = datetime.strptime(start_time_text, "%Y-%m-%d %H:%M:%S")
        finish_dt = datetime.strptime(finish_time_text, "%Y-%m-%d %H:%M:%S")

        diff = finish_dt - start_dt
        total_minutes = int(diff.total_seconds() // 60)

        if total_minutes < 0:
            return "時間異常"

        hours = total_minutes // 60
        minutes = total_minutes % 60

        return f"{hours:02d}:{minutes:02d}"

    except Exception:
        return ""


def write_trip_records_to_excel(page, records, export_path):
    """
    將出差紀錄寫入指定的 Excel 路徑。
    E欄「耗時」使用 Excel 可計算的時間格式 [h]:mm。
    """
    if not str(export_path).lower().endswith(".xlsx"):
        export_path = str(export_path) + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "出差紀錄"

    ws.append([
        "車牌",
        "目的地",
        "出差時間",
        "實際完成時間",
        "耗時",
    ])

    for plate_no, destination, out_start_time, actual_finish_time in records:
        try:
            out_start_display = datetime.strptime(
                out_start_time,
                "%Y-%m-%d %H:%M:%S"
            ).strftime("%Y/%m/%d %H:%M")
        except Exception:
            out_start_display = out_start_time

        try:
            actual_finish_display = datetime.strptime(
                actual_finish_time,
                "%Y-%m-%d %H:%M:%S"
            ).strftime("%Y/%m/%d %H:%M")
        except Exception:
            actual_finish_display = actual_finish_time

        row_no = ws.max_row + 1

        ws.cell(row=row_no, column=1, value=plate_no)
        ws.cell(row=row_no, column=2, value=destination)
        ws.cell(row=row_no, column=3, value=out_start_display)
        ws.cell(row=row_no, column=4, value=actual_finish_display)

        try:
            start_dt = datetime.strptime(out_start_time, "%Y-%m-%d %H:%M:%S")
            finish_dt = datetime.strptime(actual_finish_time, "%Y-%m-%d %H:%M:%S")

            diff = finish_dt - start_dt
            total_days = diff.total_seconds() / 86400

            if total_days >= 0:
                duration_cell = ws.cell(row=row_no, column=5, value=total_days)
                duration_cell.number_format = "[h]:mm"
            else:
                ws.cell(row=row_no, column=5, value="時間異常")

        except Exception:
            ws.cell(row=row_no, column=5, value="")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 16

    wb.save(export_path)

    show_message_dialog(
        page,
        "匯出完成",
        f"已匯出：\n{export_path}"
    )





def export_trip_records(page, start_time_text=None, end_time_text=None):
    """
    依實際完成時間區間取得資料，然後用 Windows 另存新檔視窗選擇匯出位置。
    不使用 Flet FilePicker，避免 Unknown control: FilePicker。
    """
    start_dt = parse_export_datetime(start_time_text, is_end=False)
    end_dt = parse_export_datetime(end_time_text, is_end=True)

    if start_time_text and start_dt is None:
        show_message_dialog(
            page,
            "時間格式錯誤",
            "起始時間格式錯誤，請輸入例如：2026/04/27 或 2026/04/27 08:00"
        )
        return

    if end_time_text and end_dt is None:
        show_message_dialog(
            page,
            "時間格式錯誤",
            "結束時間格式錯誤，請輸入例如：2026/04/27 或 2026/04/27 17:30"
        )
        return

    if start_dt and end_dt and start_dt > end_dt:
        show_message_dialog(page, "時間區間錯誤", "起始時間不可大於結束時間。")
        return

    start_db = start_dt.strftime("%Y-%m-%d %H:%M:%S") if start_dt else None
    end_db = end_dt.strftime("%Y-%m-%d %H:%M:%S") if end_dt else None

    records = get_trip_records_for_export(start_db, end_db)

    if not records:
        show_message_dialog(page, "無資料", "指定的實際完成時間區間內沒有出差紀錄。")
        return

    default_file_name = f"出差紀錄_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        export_path = filedialog.asksaveasfilename(
            title="選擇出差紀錄匯出位置",
            initialfile=default_file_name,
            defaultextension=".xlsx",
            filetypes=[("Excel 檔案", "*.xlsx")],
        )

        root.destroy()

        if export_path == "":
            return

        write_trip_records_to_excel(page, records, export_path)

    except PermissionError:
        show_message_dialog(
            page,
            "匯出失敗",
            "檔案可能正在被 Excel 開啟中，請先關閉後再匯出。"
        )

    except Exception as ex:
        show_message_dialog(page, "匯出失敗", str(ex))

def show_export_trip_records_dialog(page):
    today_text = datetime.now().strftime("%Y/%m/%d")

    start_text = ft.TextField(
        label="實際完成時間 - 起",
        value=today_text,
        width=320,
        hint_text="例如：2026/04/27 或 2026/04/27 08:00",
    )

    end_text = ft.TextField(
        label="實際完成時間 - 迄",
        value=today_text,
        width=320,
        hint_text="例如：2026/04/27 或 2026/04/27 17:30",
    )

    dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text("匯出出差紀錄"),
        content=ft.Column(
            controls=[
                ft.Text(
                    "請選擇要匯出的「實際完成時間」區間。",
                    size=13,
                    color=ft.Colors.GREY_700,
                ),
                start_text,
                end_text,
                ft.Text(
                    "只輸入日期時，起始會自動視為 00:00:00，結束會自動視為 23:59:59。",
                    size=12,
                    color=ft.Colors.GREY_600,
                ),
                ft.Text(
                    "按下匯出後，會跳出另存新檔視窗讓你選擇匯出位置。",
                    size=12,
                    color=ft.Colors.GREY_600,
                ),
            ],
            tight=True,
            spacing=10,
        ),
    )

    def confirm(e):
        close_dialog(page, dialog)

        export_trip_records(
            page,
            start_text.value,
            end_text.value,
        )

    def cancel(e):
        close_dialog(page, dialog)

    dialog.actions = [
        ft.TextButton("取消", on_click=cancel),
        ft.ElevatedButton("匯出", on_click=confirm),
    ]

    open_dialog(page, dialog)





ft.app(target=main)
