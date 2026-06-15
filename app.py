"""
VMS 車輛管理系統 — FastAPI Web 後端
===================================
取代原本的 Flet+Tkinter 桌面介面，提供 REST API 供前端頁面與混沌測試使用。

啟動方式：
  uvicorn app:app --reload --port 8000

前端入口：
  http://localhost:8000/
"""

from __future__ import annotations

import io
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException, Query, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

import openpyxl

from database import (
    init_db,
    seed_default_admin,
    seed_sample_data,
    reset_sample_data,
    get_vehicles,
    get_vehicle_by_plate,
    add_vehicle,
    update_vehicle,
    delete_vehicle,
    update_vehicle_status,
    update_vehicle_out,
    update_vehicle_loading,
    update_vehicle_repair,
    complete_vehicle_out,
    batch_update_vehicle_status,
    get_destinations,
    add_destination,
    update_destination,
    delete_destination,
    get_destination_minutes,
    get_loading_levels,
    add_loading_level,
    update_loading_level,
    delete_loading_level,
    authenticate_user,
    get_users,
    add_user,
    update_user,
    delete_user,
    get_all_enabled_car_nos,
    get_destination_whitelist,
    set_destination_whitelist,
    is_car_allowed_for_destination,
    get_trip_records_for_export,
)

from config import DATA_DIR, DAILY_EXPORT_DIR

# ── 資料庫初始化 ──────────────────────────────────────────────
DATA_DIR.mkdir(exist_ok=True)
DAILY_EXPORT_DIR.mkdir(exist_ok=True)
init_db()
seed_default_admin()
seed_sample_data()

# ── FastAPI App ───────────────────────────────────────────────
START_TIME = time.time()
REQUEST_METRICS: dict[tuple[str, str, str], int] = {}
REQUEST_LATENCY_SECONDS = 0.0

app = FastAPI(
    title="VMS 車輛管理系統",
    description="Vehicle Management System — REST API for Web UI & Chaos Testing",
    version="2.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# 靜態前端檔案
STATIC_DIR = Path(__file__).resolve().parent / "static"
UI_DIR = STATIC_DIR / "static" if (STATIC_DIR / "static" / "index.html").exists() else STATIC_DIR
STATIC_DIR.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(UI_DIR)), name="static")


@app.middleware("http")
async def collect_basic_metrics(request, call_next):
    global REQUEST_LATENCY_SECONDS
    started_at = time.perf_counter()
    response = await call_next(request)
    elapsed = time.perf_counter() - started_at
    REQUEST_LATENCY_SECONDS += elapsed
    if request.url.path != "/metrics":
        key = (request.method, request.url.path, str(response.status_code))
        REQUEST_METRICS[key] = REQUEST_METRICS.get(key, 0) + 1
    return response


# ── Pydantic Models ──────────────────────────────────────────
class LoginRequest(BaseModel):
    login_account: str
    password: str


class VehicleCreate(BaseModel):
    plate_no: str
    material_type: str = "A"


class VehicleUpdate(BaseModel):
    new_plate_no: str
    material_type: str


class StatusChange(BaseModel):
    status: str


class OutRequest(BaseModel):
    """出差設定 — API 直接傳入目的地與時數，不依賴 Excel。"""
    destination: str
    hours: float
    trip_time: Optional[str] = None  # "YYYY-MM-DD HH:MM:SS"，留空=now


class LoadingRequest(BaseModel):
    level: str
    hours: float


class RepairRequest(BaseModel):
    reason: str


class CompleteOutRequest(BaseModel):
    to_status: str = "standby_empty"
    actual_finish_time: Optional[str] = None  # 留空=now


class DestinationCreate(BaseModel):
    destination_name: str
    hours: float = 1.0


class DestinationUpdate(BaseModel):
    new_name: str
    hours: float


class LoadingLevelCreate(BaseModel):
    level_name: str
    hours: float = 0.5


class LoadingLevelUpdate(BaseModel):
    new_level: str
    hours: float


class UserCreate(BaseModel):
    login_account: str
    display_name: str
    password: str
    role: str = "viewer"


class UserUpdate(BaseModel):
    new_login: Optional[str] = None
    display_name: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[int] = None


class WhitelistSet(BaseModel):
    car_nos: list[str]


class BatchStatusRequest(BaseModel):
    plate_nos: list[str]
    status: str


class ChaosVehicleSeed(BaseModel):
    plate_no: str
    material_type: str = "A"
    status: str = "standby_empty"
    destination: Optional[str] = None
    hours: Optional[float] = None
    loading_level: Optional[str] = None
    repair_reason: Optional[str] = None


class ChaosSeedRequest(BaseModel):
    vehicles: list[ChaosVehicleSeed] = []
    destinations: list[DestinationCreate] = []
    loading_levels: list[LoadingLevelCreate] = []


# ── Helpers ──────────────────────────────────────────────────
def _vehicle_row_to_dict(row) -> dict:
    """将 get_vehicles() 返回的 tuple 转成 dict。"""
    return {
        "plate_no": row[0],
        "material_type": row[1],
        "status": row[2],
        "updated_at": row[3],
        "eta_time": row[4],
        "destination": row[5],
        "loading_level": row[6],
        "repair_reason": row[7],
    }


def _parse_datetime(value: Optional[str]) -> Optional[str]:
    if not value or not value.strip():
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    value = value.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(value, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    return None


# ══════════════════════════════════════════════════════════════
#   ROOT — serve index.html
# ══════════════════════════════════════════════════════════════
@app.get("/")
def root():
    index_path = UI_DIR / "index.html"
    if not index_path.exists():
        return {"message": "VMS API is running. Put index.html in /static/ to use the web UI."}
    return FileResponse(str(index_path), media_type="text/html")


@app.get("/health")
def health():
    return {
        "status": "ok",
        "service": "vms-api",
        "uptime_seconds": round(time.time() - START_TIME, 3),
        "timestamp": datetime.now().isoformat(),
    }


@app.get("/metrics")
def metrics():
    lines = [
        "# HELP vms_app_uptime_seconds Seconds since the API process started.",
        "# TYPE vms_app_uptime_seconds gauge",
        f"vms_app_uptime_seconds {time.time() - START_TIME:.3f}",
        "# HELP vms_http_requests_total Total HTTP requests handled by the API.",
        "# TYPE vms_http_requests_total counter",
    ]
    for (method, path, status), count in sorted(REQUEST_METRICS.items()):
        safe_path = path.replace("\\", "\\\\").replace('"', '\\"')
        lines.append(
            f'vms_http_requests_total{{method="{method}",path="{safe_path}",status="{status}"}} {count}'
        )
    lines.extend(
        [
            "# HELP vms_http_request_duration_seconds_total Total request handling time.",
            "# TYPE vms_http_request_duration_seconds_total counter",
            f"vms_http_request_duration_seconds_total {REQUEST_LATENCY_SECONDS:.6f}",
        ]
    )
    return Response("\n".join(lines) + "\n", media_type="text/plain; version=0.0.4")


# ══════════════════════════════════════════════════════════════
#   AUTH
# ══════════════════════════════════════════════════════════════
@app.post("/api/login")
def login(req: LoginRequest):
    user = authenticate_user(req.login_account, req.password)
    if user is None:
        raise HTTPException(status_code=401, detail="帳號或密碼錯誤，或帳號已停用。")
    return user


# ══════════════════════════════════════════════════════════════
#   VEHICLES
# ══════════════════════════════════════════════════════════════
@app.get("/api/vehicles")
def list_vehicles():
    rows = get_vehicles()
    return [_vehicle_row_to_dict(r) for r in rows]


@app.get("/api/vehicles/{plate_no}")
def get_single_vehicle(plate_no: str):
    v = get_vehicle_by_plate(plate_no)
    if v is None:
        raise HTTPException(status_code=404, detail=f"找不到車牌：{plate_no}")
    return v


@app.post("/api/vehicles", status_code=201)
def create_vehicle(req: VehicleCreate):
    try:
        add_vehicle(req.plate_no, req.material_type)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "新增成功", "plate_no": req.plate_no}


@app.put("/api/vehicles/{plate_no}")
def modify_vehicle(plate_no: str, req: VehicleUpdate):
    try:
        update_vehicle(plate_no, req.new_plate_no, req.material_type)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "更新成功"}


@app.delete("/api/vehicles/{plate_no}")
def remove_vehicle(plate_no: str):
    try:
        delete_vehicle(plate_no)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "刪除成功"}


@app.put("/api/vehicles/{plate_no}/status")
def change_vehicle_status(plate_no: str, req: StatusChange):
    """簡單狀態切換（standby_empty / standby_full）。"""
    v = get_vehicle_by_plate(plate_no)
    if v is None:
        raise HTTPException(status_code=404, detail=f"找不到車牌：{plate_no}")
    try:
        update_vehicle_status(plate_no, req.status)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "狀態已更新", "plate_no": plate_no, "status": req.status}


@app.put("/api/vehicles/{plate_no}/out")
def vehicle_out(plate_no: str, req: OutRequest):
    """出差設定 — 直接用 API JSON 傳入資料，不依賴 Excel。"""
    v = get_vehicle_by_plate(plate_no)
    if v is None:
        raise HTTPException(status_code=404, detail=f"找不到車牌：{plate_no}")

    # 白名單檢查
    if not is_car_allowed_for_destination(req.destination, plate_no):
        raise HTTPException(
            status_code=403,
            detail=f"非白名單：車牌 {plate_no} 未被允許前往「{req.destination}」",
        )

    trip_time = _parse_datetime(req.trip_time)
    if trip_time is None:
        raise HTTPException(status_code=400, detail="trip_time 格式錯誤")

    try:
        update_vehicle_out(plate_no, req.destination, req.hours, trip_time)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    return {
        "message": "已設定出差",
        "plate_no": plate_no,
        "destination": req.destination,
        "hours": req.hours,
    }


@app.put("/api/vehicles/{plate_no}/loading")
def vehicle_loading(plate_no: str, req: LoadingRequest):
    v = get_vehicle_by_plate(plate_no)
    if v is None:
        raise HTTPException(status_code=404, detail=f"找不到車牌：{plate_no}")
    loading_minutes = req.hours * 60
    try:
        update_vehicle_loading(plate_no, req.level, loading_minutes)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "已設定裝料", "plate_no": plate_no, "level": req.level}


@app.put("/api/vehicles/{plate_no}/repair")
def vehicle_repair(plate_no: str, req: RepairRequest):
    v = get_vehicle_by_plate(plate_no)
    if v is None:
        raise HTTPException(status_code=404, detail=f"找不到車牌：{plate_no}")
    try:
        update_vehicle_repair(plate_no, req.reason)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "已設定修車", "plate_no": plate_no}


@app.put("/api/vehicles/{plate_no}/complete-out")
def vehicle_complete_out(plate_no: str, req: CompleteOutRequest):
    v = get_vehicle_by_plate(plate_no)
    if v is None:
        raise HTTPException(status_code=404, detail=f"找不到車牌：{plate_no}")

    finish_time = _parse_datetime(req.actual_finish_time)
    if finish_time is None:
        raise HTTPException(status_code=400, detail="actual_finish_time 格式錯誤")

    try:
        complete_vehicle_out(plate_no, req.to_status, finish_time)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    return {"message": "出差已完成", "plate_no": plate_no, "to_status": req.to_status}


# ══════════════════════════════════════════════════════════════
#   DESTINATIONS
# ══════════════════════════════════════════════════════════════
@app.get("/api/destinations")
def list_destinations():
    rows = get_destinations()
    return [{"destination_name": r[0], "hours": r[1]} for r in rows]


@app.post("/api/destinations", status_code=201)
def create_destination(req: DestinationCreate):
    try:
        add_destination(req.destination_name, req.hours)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "新增成功"}


@app.put("/api/destinations/{name}")
def modify_destination(name: str, req: DestinationUpdate):
    try:
        update_destination(name, req.new_name, req.hours)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "更新成功"}


@app.delete("/api/destinations/{name}")
def remove_destination(name: str):
    try:
        delete_destination(name)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "刪除成功"}


# ══════════════════════════════════════════════════════════════
#   LOADING LEVELS
# ══════════════════════════════════════════════════════════════
@app.get("/api/loading-levels")
def list_loading_levels():
    rows = get_loading_levels()
    return [{"level_name": r[0], "hours": r[1]} for r in rows]


@app.post("/api/loading-levels", status_code=201)
def create_loading_level(req: LoadingLevelCreate):
    try:
        add_loading_level(req.level_name, req.hours)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "新增成功"}


@app.put("/api/loading-levels/{name}")
def modify_loading_level(name: str, req: LoadingLevelUpdate):
    try:
        update_loading_level(name, req.new_level, req.hours)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "更新成功"}


@app.delete("/api/loading-levels/{name}")
def remove_loading_level(name: str):
    try:
        delete_loading_level(name)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "刪除成功"}


# ══════════════════════════════════════════════════════════════
#   USERS
# ══════════════════════════════════════════════════════════════
@app.get("/api/users")
def list_users():
    rows = get_users()
    return [
        {"login_account": r[0], "display_name": r[1], "role": r[2], "is_active": r[3]}
        for r in rows
    ]


@app.post("/api/users", status_code=201)
def create_user(req: UserCreate):
    try:
        add_user(req.login_account, req.display_name, req.password, req.role)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "新增成功"}


@app.put("/api/users/{account}")
def modify_user(account: str, req: UserUpdate):
    try:
        update_user(
            account,
            req.new_login or account,
            req.display_name or "",
            req.password or "",
            req.role or "viewer",
            req.is_active if req.is_active is not None else 1,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "更新成功"}


@app.delete("/api/users/{account}")
def remove_user(account: str):
    try:
        delete_user(account)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"message": "刪除成功"}


# ══════════════════════════════════════════════════════════════
#   WHITELIST
# ══════════════════════════════════════════════════════════════
@app.get("/api/whitelist/{destination}")
def get_whitelist(destination: str):
    cars = get_destination_whitelist(destination)
    all_cars = get_all_enabled_car_nos()
    return {"destination": destination, "selected": cars, "all_cars": all_cars}


@app.put("/api/whitelist/{destination}")
def set_whitelist(destination: str, req: WhitelistSet):
    set_destination_whitelist(destination, req.car_nos)
    return {"message": "白名單已儲存"}


# ══════════════════════════════════════════════════════════════
#   TRIP RECORDS
# ══════════════════════════════════════════════════════════════
@app.get("/api/trip-records")
def list_trip_records(
    start: Optional[str] = Query(None),
    end: Optional[str] = Query(None),
):
    start_db = _parse_datetime(start) if start else None
    end_db = _parse_datetime(end) if end else None
    rows = get_trip_records_for_export(start_db, end_db)
    return [
        {
            "plate_no": r[0],
            "destination": r[1],
            "out_start_time": r[2],
            "actual_finish_time": r[3],
        }
        for r in rows
    ]


@app.get("/api/trip-records/export")
def export_trip_records_excel(
    start: Optional[str] = Query(None),
    end: Optional[str] = Query(None),
):
    start_db = _parse_datetime(start) if start else None
    end_db = _parse_datetime(end) if end else None
    records = get_trip_records_for_export(start_db, end_db)

    if not records:
        raise HTTPException(status_code=404, detail="指定區間內無出差紀錄。")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "出差紀錄"
    ws.append(["車牌", "目的地", "出差時間", "實際完成時間", "耗時"])

    for plate_no, destination, out_start_time, actual_finish_time in records:
        row_no = ws.max_row + 1
        ws.cell(row=row_no, column=1, value=plate_no)
        ws.cell(row=row_no, column=2, value=destination)
        ws.cell(row=row_no, column=3, value=out_start_time)
        ws.cell(row=row_no, column=4, value=actual_finish_time)

        try:
            s = datetime.strptime(out_start_time, "%Y-%m-%d %H:%M:%S")
            f = datetime.strptime(actual_finish_time, "%Y-%m-%d %H:%M:%S")
            diff_days = (f - s).total_seconds() / 86400
            if diff_days >= 0:
                cell = ws.cell(row=row_no, column=5, value=diff_days)
                cell.number_format = "[h]:mm"
            else:
                ws.cell(row=row_no, column=5, value="時間異常")
        except Exception:
            ws.cell(row=row_no, column=5, value="")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"trip_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ══════════════════════════════════════════════════════════════
#   CHAOS TESTING — No Auth
# ══════════════════════════════════════════════════════════════
@app.get("/api/chaos/health")
def chaos_health():
    """健康檢查。"""
    vehicles = get_vehicles()
    status_counts = {}
    for r in vehicles:
        s = r[2]
        status_counts[s] = status_counts.get(s, 0) + 1
    return {
        "status": "ok",
        "timestamp": datetime.now().isoformat(),
        "total_vehicles": len(vehicles),
        "status_counts": status_counts,
    }


@app.post("/api/chaos/reset")
def chaos_reset():
    """重設資料庫為預設樣本資料。"""
    reset_sample_data()
    seed_default_admin()
    return {"message": "資料庫已重設為預設樣本資料", "timestamp": datetime.now().isoformat()}


@app.post("/api/chaos/seed")
def chaos_seed(req: ChaosSeedRequest):
    """注入自定義測試資料。"""
    results = {"vehicles_added": 0, "destinations_added": 0, "loading_levels_added": 0, "errors": []}

    for v in req.vehicles:
        try:
            add_vehicle(v.plate_no, v.material_type)
            if v.status == "out" and v.destination and v.hours:
                update_vehicle_out(v.plate_no, v.destination, v.hours)
            elif v.status == "loading" and v.loading_level and v.hours:
                update_vehicle_loading(v.plate_no, v.loading_level, v.hours * 60)
            elif v.status == "repair" and v.repair_reason:
                update_vehicle_repair(v.plate_no, v.repair_reason)
            elif v.status not in ("standby_empty",):
                update_vehicle_status(v.plate_no, v.status)
            results["vehicles_added"] += 1
        except Exception as e:
            results["errors"].append(f"vehicle {v.plate_no}: {e}")

    for d in req.destinations:
        try:
            add_destination(d.destination_name, d.hours)
            results["destinations_added"] += 1
        except Exception as e:
            results["errors"].append(f"destination {d.destination_name}: {e}")

    for ll in req.loading_levels:
        try:
            add_loading_level(ll.level_name, ll.hours)
            results["loading_levels_added"] += 1
        except Exception as e:
            results["errors"].append(f"loading_level {ll.level_name}: {e}")

    return results


@app.post("/api/chaos/batch-status")
def chaos_batch_status(req: BatchStatusRequest):
    """批次更改多車狀態（壓力測試用）。"""
    count = batch_update_vehicle_status(req.plate_nos, req.status)
    return {"message": f"已更新 {count} 台車輛", "updated": count}


# ══════════════════════════════════════════════════════════════
#   MAIN — 直接 python app.py 也可以啟動
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
